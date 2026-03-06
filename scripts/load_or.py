#!/usr/bin/env python3
"""
Oregon data loader: hunts, GMUs, draw results, hunt dates.

Sources:
  - 2024/2025 preference point draw reports (elk + buck deer)
  - OR/proclamations/2026/OR_hunt_dates_2026.csv
"""

import os
import re
import csv
import openpyxl
import psycopg2

BASE_DIR = "/Users/openclaw/Documents/GraysonsDrawOdds"
DB_CONFIG = {
    'host': 'localhost', 'port': 5432,
    'dbname': 'draws', 'user': 'draws', 'password': 'drawspass'
}

# Weapon-type words to strip when extracting GMU name from Hunt Name
STRIP_WORDS = [
    'Muzzleloader', 'Archery', 'Rifle', 'Bull', 'Cow', 'Buck', 'Doe',
    'Youth', 'Premium', 'Either Sex', 'Antlered', 'Antlerless', 'Any',
    'Legal Weapon', 'Special', 'General', 'Bow', 'Oregon'
]
STRIP_RE = re.compile(
    r'\b(?:' + '|'.join(re.escape(w) for w in STRIP_WORDS) + r')\b',
    re.IGNORECASE
)


def parse_weapon_type(hunt_code, hunt_name):
    """Return weapon_type_id based on hunt code suffix and name."""
    code = str(hunt_code).strip().upper()
    name = (hunt_name or '').lower()
    if code.endswith('M') or 'muzzleloader' in name:
        return 4  # MUZZ
    if code.endswith('A') or 'archery' in name or 'bow' in name:
        return 3  # ARCHERY
    return 1  # ANY (Oregon "rifle" hunts are really "any legal weapon")


def parse_sex(hunt_name, species_code):
    """Return bag_limit_id from hunt name context."""
    name = (hunt_name or '').lower()
    if species_code == 'ELK':
        if 'cow' in name or 'antlerless' in name:
            return 15  # COW
        if 'bull' in name or 'antlered' in name:
            return 13  # BULL
        if 'either sex' in name or 'any' in name:
            return 5   # ES (either sex)
        if 'spike' in name:
            return 14  # SPIKE
        return 13  # default elk = BULL
    else:  # MDR
        if 'doe' in name or 'antlerless' in name:
            return 18  # DOE
        if 'buck' in name or 'antlered' in name:
            return 16  # BUCK
        if 'either sex' in name or 'any' in name:
            return 5   # ES
        return 16  # default deer = BUCK


def extract_gmu_name(hunt_name):
    """Strip weapon/sex words from hunt name to get GMU/unit name."""
    name = STRIP_RE.sub('', hunt_name or '')
    name = re.sub(r'\s+', ' ', name).strip()
    return name if name else hunt_name


def parse_2024_file(filepath):
    """Parse 2024-format draw report: row 5 = headers, row 6+ = data.
    Only read hunt-level summary rows (hunt number is not None)."""
    wb = openpyxl.load_workbook(filepath, read_only=True)
    ws = wb.active
    hunts = []
    for row in ws.iter_rows(min_row=6, values_only=True):
        hunt_num = row[0]
        if hunt_num is None or str(hunt_num).strip() == '':
            continue
        hunts.append({
            'hunt_code': str(hunt_num).strip(),
            'hunt_name': str(row[1]).strip() if row[1] else '',
            'tags_authorized': int(row[2]) if row[2] else 0,
            'res_apps': int(row[3]) if row[3] else 0,
            'res_drawn': int(row[4]) if row[4] else 0,
            'nr_apps': int(row[5]) if row[5] else 0,
            'nr_drawn': int(row[6]) if row[6] else 0,
            'total_apps': int(row[7]) if row[7] else 0,
            'total_drawn': int(row[8]) if row[8] else 0,
            'pts_apps': int(row[9]) if row[9] else 0,
            'pts_drawn_p1': int(row[10]) if row[10] else 0,
            'pts_drawn_p2': int(row[11]) if row[11] else 0,
        })
    wb.close()
    return hunts


def parse_2025_file(filepath):
    """Parse 2025-format draw report: row 1 = headers, row 2+ = data.
    Hunt info is repeated on every row; aggregate by hunt_code."""
    wb = openpyxl.load_workbook(filepath, read_only=True)
    ws = wb.active
    hunts = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        hunt_num = row[0]
        if hunt_num is None or str(hunt_num).strip() == '':
            continue
        code = str(hunt_num).strip()
        if code not in hunts:
            hunts[code] = {
                'hunt_code': code,
                'hunt_name': str(row[1]).strip() if row[1] else '',
                'tags_authorized': int(row[2]) if row[2] else 0,
                'res_apps': int(row[3]) if row[3] else 0,
                'res_drawn': int(row[4]) if row[4] else 0,
                'nr_apps': int(row[5]) if row[5] else 0,
                'nr_drawn': int(row[6]) if row[6] else 0,
                'total_apps': int(row[7]) if row[7] else 0,
                'total_drawn': int(row[8]) if row[8] else 0,
                'pts_apps': int(row[9]) if row[9] else 0,
                'pts_drawn_p1': int(row[10]) if row[10] else 0,
                'pts_drawn_p2': int(row[11]) if row[11] else 0,
            }
    wb.close()
    return list(hunts.values())


def gmu_code_from_hunt(hunt_code):
    """Extract numeric prefix as GMU code (first 3 digits for elk 200-series,
    first 3 for deer 100-series)."""
    m = re.match(r'^(\d+)', hunt_code)
    return m.group(1) if m else hunt_code


def main():
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()

    # Get state_id
    cur.execute("SELECT state_id FROM states WHERE state_code='OR'")
    or_state_id = cur.fetchone()[0]

    # Species map
    cur.execute("SELECT species_id, species_code FROM species")
    species_map = {r[1]: r[0] for r in cur.fetchall()}

    # Ensure pools exist for OR
    for pool_code, desc, pct, note in [
        ('RES', 'Resident pool', 95.0, '95% of tags'),
        ('NR', 'Nonresident pool', 5.0, '5% of tags'),
    ]:
        cur.execute("""
            INSERT INTO pools (state_id, pool_code, description, allocation_pct, allocation_note)
            VALUES (%s, %s, %s, %s, %s)
            ON CONFLICT (state_id, pool_code) DO NOTHING
        """, (or_state_id, pool_code, desc, pct, note))

    # Get pool IDs
    cur.execute("SELECT pool_id, pool_code FROM pools WHERE state_id = %s", (or_state_id,))
    pool_map = {r[1]: r[0] for r in cur.fetchall()}
    conn.commit()

    # Define source files: (filepath, year, species_code, file_format)
    sources = [
        ('OR/raw_data/2024_elk_preference_point_draw_report.xlsx', 2024, 'ELK', '2024'),
        ('OR/raw_data/2024_buck_deer_preference_point_draw_report.xlsx', 2024, 'MDR', '2024'),
        ('OR/raw_data/2025_elk_preference_point_draw_report.xlsx', 2025, 'ELK', '2025'),
        ('OR/raw_data/2025_buck_deer_preference_point_draw_report.xlsx', 2025, 'MDR', '2025'),
    ]

    total_hunts = 0
    total_gmus = 0
    total_draw = 0

    for relpath, year, sp_code, fmt in sources:
        filepath = os.path.join(BASE_DIR, relpath)
        if not os.path.exists(filepath):
            print(f"  SKIP (missing): {relpath}")
            continue

        species_id = species_map[sp_code]
        if fmt == '2024':
            hunts = parse_2024_file(filepath)
        else:
            hunts = parse_2025_file(filepath)

        print(f"\n  {relpath}: {len(hunts)} hunts parsed")

        for h in hunts:
            hunt_code = h['hunt_code']
            hunt_name = h['hunt_name']
            weapon_type_id = parse_weapon_type(hunt_code, hunt_name)
            bag_limit_id = parse_sex(hunt_name, sp_code)
            gmu_name = extract_gmu_name(hunt_name)
            gmu_code = gmu_code_from_hunt(hunt_code)
            gmu_sort_key = gmu_code.zfill(5)

            # Determine tag_type / season_type / notes
            notes = None
            season_type = 'controlled'
            tag_type = 'LE'
            if 'youth' in hunt_name.lower():
                notes = 'Youth hunt'
            if 'premium' in hunt_name.lower():
                notes = 'Premium draw'
                season_type = 'premium'

            # Insert GMU
            cur.execute("""
                INSERT INTO gmus (state_id, gmu_code, gmu_name, gmu_sort_key)
                VALUES (%s, %s, %s, %s)
                ON CONFLICT (state_id, gmu_code) DO UPDATE SET gmu_name = EXCLUDED.gmu_name
                RETURNING gmu_id
            """, (or_state_id, gmu_code, gmu_name, gmu_sort_key))
            gmu_id = cur.fetchone()[0]
            total_gmus += 1

            # Insert hunt
            cur.execute("""
                INSERT INTO hunts (state_id, species_id, hunt_code, hunt_code_display,
                    weapon_type_id, bag_limit_id, season_type, tag_type, is_active,
                    unit_description, notes)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, 1, %s, %s)
                ON CONFLICT (state_id, hunt_code) DO UPDATE SET
                    weapon_type_id = EXCLUDED.weapon_type_id,
                    bag_limit_id = EXCLUDED.bag_limit_id,
                    season_type = EXCLUDED.season_type,
                    unit_description = EXCLUDED.unit_description,
                    notes = EXCLUDED.notes
                RETURNING hunt_id
            """, (or_state_id, species_id, hunt_code, hunt_code,
                  weapon_type_id, bag_limit_id, season_type, tag_type,
                  gmu_name, notes))
            hunt_id = cur.fetchone()[0]
            total_hunts += 1

            # Link hunt to GMU
            cur.execute("""
                INSERT INTO hunt_gmus (hunt_id, gmu_id) VALUES (%s, %s)
                ON CONFLICT (hunt_id, gmu_id) DO NOTHING
            """, (hunt_id, gmu_id))

            # Insert draw results by pool
            for pool_code, apps, drawn in [
                ('RES', h['res_apps'], h['res_drawn']),
                ('NR', h['nr_apps'], h['nr_drawn']),
            ]:
                if apps == 0 and drawn == 0:
                    continue
                pool_id = pool_map[pool_code]
                odds = round(drawn / apps, 6) if apps > 0 else 0.0
                # min_pts_drawn: use pts_drawn_p1 as proxy (pref-point round drawn count)
                min_pts = h['pts_drawn_p1'] if h['pts_drawn_p1'] > 0 else None

                cur.execute("""
                    INSERT INTO draw_results_by_pool
                        (hunt_id, draw_year, pool_id, applications, tags_available,
                         tags_awarded, min_pts_drawn)
                    VALUES (%s, %s, %s, %s, %s, %s, %s)
                    ON CONFLICT (hunt_id, draw_year, pool_id) DO UPDATE SET
                        applications = EXCLUDED.applications,
                        tags_available = EXCLUDED.tags_available,
                        tags_awarded = EXCLUDED.tags_awarded,
                        min_pts_drawn = EXCLUDED.min_pts_drawn
                """, (hunt_id, year, pool_id, apps, h['tags_authorized'],
                      drawn, min_pts))
                total_draw += 1

        conn.commit()

    # Load hunt dates
    dates_csv = os.path.join(BASE_DIR, 'OR/proclamations/2026/OR_hunt_dates_2026.csv')
    dates_loaded = 0
    dates_unmatched = 0
    if os.path.exists(dates_csv):
        cur.execute("SELECT hunt_code, hunt_id FROM hunts WHERE state_id = %s", (or_state_id,))
        hunt_map = {r[0]: r[1] for r in cur.fetchall()}
        with open(dates_csv) as f:
            reader = csv.DictReader(f)
            for row in reader:
                hc = row['hunt_code'].strip()
                if hc in hunt_map:
                    cur.execute("""
                        INSERT INTO hunt_dates (hunt_id, season_year, start_date, end_date, notes)
                        VALUES (%s, 2026, %s, %s, %s)
                        ON CONFLICT (hunt_id, season_year) DO UPDATE SET
                            start_date = EXCLUDED.start_date,
                            end_date = EXCLUDED.end_date,
                            notes = EXCLUDED.notes
                    """, (hunt_map[hc], row['open_date'], row['close_date'],
                          row.get('notes', '')))
                    dates_loaded += 1
                else:
                    dates_unmatched += 1
        conn.commit()

    # Print counts
    print("\n=== OR LOAD SUMMARY ===")
    cur.execute("SELECT COUNT(*) FROM hunts WHERE state_id = %s", (or_state_id,))
    print(f"  Hunts:        {cur.fetchone()[0]}")
    cur.execute("SELECT COUNT(*) FROM gmus WHERE state_id = %s", (or_state_id,))
    print(f"  GMUs:         {cur.fetchone()[0]}")
    cur.execute("""SELECT COUNT(*) FROM draw_results_by_pool dr
                   JOIN hunts h ON h.hunt_id = dr.hunt_id WHERE h.state_id = %s""", (or_state_id,))
    print(f"  Draw results: {cur.fetchone()[0]}")
    cur.execute("""SELECT COUNT(*) FROM hunt_dates hd
                   JOIN hunts h ON h.hunt_id = hd.hunt_id WHERE h.state_id = %s""", (or_state_id,))
    print(f"  Hunt dates:   {cur.fetchone()[0]}")
    print(f"  Dates loaded: {dates_loaded}, unmatched: {dates_unmatched}")

    conn.close()
    print("\nOR load complete.")


if __name__ == '__main__':
    main()
